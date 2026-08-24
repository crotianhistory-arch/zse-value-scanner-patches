from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import load_workbook


class AdapterError(RuntimeError):
    pass


def _h(v: Any) -> str:
    return " ".join(
        re.sub(r"[^a-z0-9]+", " ", str(v or "").casefold()).split()
    )


def _code(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    m = re.fullmatch(
        r"[A-Za-z]?([0-9]{1,4})",
        str(v).strip().replace(" ", "").replace("'", ""),
    )
    return m.group(1).zfill(4) if m else None


def _revision_header(text: str, revision: int) -> bool:
    return any(
        marker in text
        for marker in (
            f"rev {revision}",
            f"rev{revision}",
            f"revision {revision}",
            f"revision{revision}",
        )
    )


def _pick_code_column(
    ws: Any,
    candidates: list[int],
    header_text: dict[int, str],
    data_start: int,
) -> int:
    """Choose the four-digit class-code column, not an alphanumerical helper code."""
    sample_end = min(ws.max_row, data_start + 250)

    def score(column: int) -> tuple[int, int, int, int]:
        text = header_text[column]
        alpha_helper = int(
            "alphanumeric" in text or "alphanumerical" in text
        )
        numeric_hint = int(
            "numeric" in text or "numerical" in text
        )
        parseable = sum(
            _code(ws.cell(row, column).value) is not None
            for row in range(data_start, sample_end + 1)
        )
        # Prefer non-alphanumerical columns, then explicitly numerical ones,
        # then the column that actually parses the most class codes.
        return (-alpha_helper, numeric_hint, parseable, -column)

    return max(candidates, key=score)


def _headers(
    ws: Any,
) -> tuple[int, int, int, int | None, int | None] | None:
    # Official statistical workbooks often use merged or multi-row headers.
    # Search 1-3 adjacent rows and combine the visible text by column rather
    # than baking one workbook layout into the adapter.
    max_row = min(ws.max_row, 30)
    max_col = min(ws.max_column, 40)

    for start in range(1, max_row + 1):
        for span in (1, 2, 3):
            if start + span - 1 > max_row:
                continue

            vals: dict[int, str] = {}
            for column in range(1, max_col + 1):
                parts = [
                    _h(ws.cell(row, column).value)
                    for row in range(start, start + span)
                ]
                vals[column] = " ".join(x for x in parts if x)

            rev4 = [
                column
                for column, text in vals.items()
                if "isic" in text
                and _revision_header(text, 4)
                and "code" in text
                and "title" not in text
            ]
            rev5 = [
                column
                for column, text in vals.items()
                if "isic" in text
                and _revision_header(text, 5)
                and "code" in text
                and "title" not in text
            ]

            if not rev4 or not rev5:
                continue

            data_start = start + span
            from_col = _pick_code_column(
                ws, rev4, vals, data_start
            )
            to_col = _pick_code_column(
                ws, rev5, vals, data_start
            )

            change = next(
                (
                    column
                    for column, text in vals.items()
                    if ("change" in text or "gsim" in text)
                    and ("type" in text or "category" in text)
                ),
                None,
            )
            note = next(
                (
                    column
                    for column, text in vals.items()
                    if ("description" in text or "note" in text)
                    and ("change" in text or "content" in text)
                ),
                None,
            )

            return (
                data_start,
                from_col,
                to_col,
                change,
                note,
            )

    return None


def parse_unsd_isic_rev4_rev5_xlsx(
    data: bytes,
) -> list[dict[str, Any]]:
    try:
        wb = load_workbook(
            io.BytesIO(data),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise AdapterError(
            f"invalid UNSD workbook: {exc}"
        ) from exc

    candidates: list[list[dict[str, Any]]] = []

    for ws in wb.worksheets:
        header = _headers(ws)
        if not header:
            continue

        start, from_col, to_col, change_col, note_col = header
        rows: list[dict[str, Any]] = []
        seen: set[tuple[str, str, str | None, str | None]] = set()
        blanks = 0

        for row in range(start, ws.max_row + 1):
            from_code = _code(
                ws.cell(row, from_col).value
            )
            to_code = _code(
                ws.cell(row, to_col).value
            )

            if not from_code and not to_code:
                blanks += 1
                if blanks >= 25 and rows:
                    break
                continue

            blanks = 0
            if not from_code or not to_code:
                continue

            change = (
                " ".join(
                    str(ws.cell(row, change_col).value).split()
                )
                if change_col
                and ws.cell(row, change_col).value is not None
                else None
            )
            note = (
                " ".join(
                    str(ws.cell(row, note_col).value).split()
                )
                if note_col
                and ws.cell(row, note_col).value is not None
                else None
            )

            key = (
                from_code,
                to_code,
                change,
                note,
            )
            if key in seen:
                continue

            seen.add(key)
            rows.append(
                {
                    "from_code": from_code,
                    "to_code": to_code,
                    "official_change_type": change,
                    "official_note": note,
                }
            )

        if rows:
            candidates.append(rows)

    if not candidates:
        raise AdapterError(
            "ISIC Rev.4/Rev.5 table not found in workbook"
        )

    return max(
        candidates,
        key=lambda rows: (
            len({x["from_code"] for x in rows}),
            len({x["to_code"] for x in rows}),
            len(rows),
        ),
    )

from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook

from zse_tool.classification_crosswalk_unsd import (
    _headers,
    parse_unsd_isic_rev4_rev5_xlsx,
)


def _real_unsd_shape_fixture() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "ISIC4-5"
    ws.append([
        "ISIC Rev. 4 Alphanumerical Code",
        "ISIC Rev. 4 Code",
        "ISIC Rev. 4 Title",
        "ISIC Rev. 5 Alphanumerical Code",
        "ISIC Rev. 5 Code",
        "ISIC Rev. 5 Title",
        "GSIM Type of Change",
        "Description of changed content",
    ])
    ws.append([
        "A0111",
        "0111",
        "Growing of cereals (except rice), leguminous crops and oil seeds",
        "A0111",
        "0111",
        "Growing of cereals (except rice), leguminous crops and oil seeds",
        "RC1 - No change",
        None,
    ])
    ws.append([
        "A013A0",
        "0130",
        "Plant propagation",
        "A0130",
        "0130",
        "Plant propagation",
        "VC1 - Code change",
        "Alphanumerical code error corrected.",
    ])
    ws.append([
        "A017A0",
        "0170",
        "Hunting, trapping and related service activities",
        "A0170",
        "0170",
        "Hunting, trapping and related service activities",
        "VC1 - Code change",
        "Alphanumerical code error corrected.",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_unsd_adapter_prefers_numeric_class_code_columns():
    data = _real_unsd_shape_fixture()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    header = _headers(wb.active)
    assert header is not None
    _, from_col, to_col, _, _ = header
    assert from_col == 2
    assert to_col == 5

    rows = parse_unsd_isic_rev4_rev5_xlsx(data)
    assert {(row["from_code"], row["to_code"]) for row in rows} == {
        ("0111", "0111"),
        ("0130", "0130"),
        ("0170", "0170"),
    }
    corrected = {
        row["from_code"]: row["official_note"]
        for row in rows
        if row["from_code"] in {"0130", "0170"}
    }
    assert corrected == {
        "0130": "Alphanumerical code error corrected.",
        "0170": "Alphanumerical code error corrected.",
    }
